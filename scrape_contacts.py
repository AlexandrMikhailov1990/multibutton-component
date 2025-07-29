#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import requests
import re
import time
from urllib.parse import urljoin, urlparse
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import urllib3

# Отключаем предупреждения о SSL
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

def clean_url(url):
    """Очищает и нормализует URL"""
    if not url.startswith(('http://', 'https://')):
        url = 'https://' + url
    return url

def extract_phone_numbers(text):
    """Извлекает номера телефонов из текста"""
    # Паттерны для российских номеров
    patterns = [
        r'8\s*\(?800\)?\s*\d{3}[-\s]?\d{2}[-\s]?\d{2}',  # 8 (800) XXX-XX-XX
        r'8\s*\(?495\)?\s*\d{3}[-\s]?\d{2}[-\s]?\d{2}',  # 8 (495) XXX-XX-XX
        r'\+7\s*\(?800\)?\s*\d{3}[-\s]?\d{2}[-\s]?\d{2}', # +7 (800) XXX-XX-XX
        r'\+7\s*\(?495\)?\s*\d{3}[-\s]?\d{2}[-\s]?\d{2}', # +7 (495) XXX-XX-XX
        r'8\s*\(?[0-9]{3}\)?\s*\d{3}[-\s]?\d{2}[-\s]?\d{2}', # 8 (XXX) XXX-XX-XX
        r'\+7\s*\(?[0-9]{3}\)?\s*\d{3}[-\s]?\d{2}[-\s]?\d{2}', # +7 (XXX) XXX-XX-XX
        r'8\s*\d{10}',  # 8 XXXXXXXXXX
        r'\+7\s*\d{10}',  # +7 XXXXXXXXXX
    ]
    
    phones = []
    for pattern in patterns:
        matches = re.findall(pattern, text, re.IGNORECASE)
        phones.extend(matches)
    
    return list(set(phones))  # Убираем дубликаты

def extract_emails(text):
    """Извлекает email адреса из текста"""
    email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
    emails = re.findall(email_pattern, text, re.IGNORECASE)
    return list(set(emails))  # Убираем дубликаты

def scrape_website(url, company_name):
    """Собирает контактную информацию с сайта"""
    try:
        # Очищаем URL
        clean_url_str = clean_url(url)
        
        # Настройки запроса
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        
        # Пробуем сначала с HTTPS, потом с HTTP
        for protocol in ['https', 'http']:
            try:
                test_url = f"{protocol}://{url.replace('https://', '').replace('http://', '')}"
                response = requests.get(test_url, headers=headers, timeout=10, verify=False)
                response.raise_for_status()
                clean_url_str = test_url
                break
            except:
                continue
        else:
            # Если не удалось подключиться, пробуем оригинальный URL
            response = requests.get(clean_url_str, headers=headers, timeout=10, verify=False)
            response.raise_for_status()
        
        # Парсим HTML
        soup = BeautifulSoup(response.text, 'html.parser')
        
        # Получаем весь текст
        text = soup.get_text()
        
        # Извлекаем телефоны и email
        phones = extract_phone_numbers(text)
        emails = extract_emails(text)
        
        # Ищем контактную информацию в ссылках
        contact_links = soup.find_all('a', href=True)
        for link in contact_links:
            href = link.get('href', '')
            link_text = link.get_text()
            
            # Проверяем ссылки на контакты
            if any(keyword in href.lower() for keyword in ['contact', 'kontakt', 'kontakty', 'about', 'o-nas', 'kontakt']):
                try:
                    contact_url = urljoin(clean_url_str, href)
                    contact_response = requests.get(contact_url, headers=headers, timeout=5, verify=False)
                    if contact_response.status_code == 200:
                        contact_soup = BeautifulSoup(contact_response.text, 'html.parser')
                        contact_text = contact_soup.get_text()
                        phones.extend(extract_phone_numbers(contact_text))
                        emails.extend(extract_emails(contact_text))
                except:
                    pass
            
            # Проверяем mailto ссылки
            if href.startswith('mailto:'):
                email = href.replace('mailto:', '').split('?')[0]
                if email not in emails:
                    emails.append(email)
        
        # Убираем дубликаты
        phones = list(set(phones))
        emails = list(set(emails))
        
        # Фильтруем и форматируем телефоны
        filtered_phones = []
        for phone in phones:
            # Убираем лишние символы и форматируем
            phone_clean = re.sub(r'[^\d+]', '', phone)
            if len(phone_clean) >= 10:  # Минимальная длина для телефона
                filtered_phones.append(phone)
        
        # Форматируем результат
        phone_result = filtered_phones[0] if filtered_phones else "Не найден"
        email_result = emails[0] if emails else "Не найден"
        
        print(f"✓ {company_name}: Телефон - {phone_result}, Email - {email_result}")
        return phone_result, email_result
        
    except Exception as e:
        print(f"✗ {company_name}: Ошибка - {str(e)}")
        return "Ошибка загрузки", "Ошибка загрузки"

def update_excel_with_contacts():
    """Обновляет Excel файл с найденными контактами"""
    # Открываем существующий файл
    wb = openpyxl.load_workbook('companies_data.xlsx')
    
    # Получаем лист с общими компаниями
    ws = wb['Общие компании']
    
    # Данные компаний для обновления
    companies_to_update = [
        ["RN-Card", "rn-card.ru"],
        ["DHL Россия", "zakaz.dhl.ru"],
        ["УГМК-Клиника", "ugmk-clinic.ru"],
        ["RegionSale", "regionsale.ru"],
        ["Радуга Камня", "radugakamnya.ru"],
        ["CopyTimer", "copytimer.ur.ru"],
        ["Lesk", "lesk.ru"],
        ["РГДБ", "rgdb.ru"],
        ["Lady Maria", "lady-maria.ru"],
        ["UniCredit Bank", "unicreditbank.ru"],
        ["Делікатеска", "delikateska.ru"],
        ["Russian Flower", "russianflower.ru"],
        ["Ростсельмаш", "rostselmash.com"],
        ["DAOffice", "daoffice.ru"],
        ["Felix", "felix1.dterra.ru"],
        ["Linline Clinic", "linline-clinic.ru"],
        ["Unetcom", "unetcom.ru"],
        ["Декарт", "dekart.ru"],
        ["Homework", "homework.ru"],
        ["HayPost", "haypost.am"],
        ["R-Seven", "r-seven.ru"],
        ["Рыболов", "rybolov.org"],
        ["Deuter Shop", "deuter-shop.ru"],
        ["Зарождение", "zarozhdenie43.ru"],
        ["Dareco", "dareco.ru"],
        ["Dezar", "dezar.su"],
        ["Diplomart", "diplomart.ru"],
        ["Дом Магии", "dommagii.com"],
        ["Доступ Окна", "dostupokna.ru"],
        ["Dvaris", "dvaris.ru"],
        ["Hazina Tur", "hazinatur.ru"],
        ["Labcentrifuge", "labcentrifuge.ru"],
        ["Лестницы Просто", "lestnicy-prosto.ru"],
        ["LiveTex", "livetex.ru"],
        ["Luki v Ruki", "lukivruki.ru"],
        ["Reglet", "reglet.ru"],
        ["Renault Avangard", "renault-avangard.ru"],
        ["Roskedr", "roskedr.ru"],
        ["RPBeton", "rpbeton.ru"],
        ["RTG-MPS", "rtg-mps.ru"]
    ]
    
    print("Начинаю сбор контактной информации...")
    print("=" * 50)
    
    # Обновляем данные в Excel
    for row_idx, (company_name, website) in enumerate(companies_to_update, start=2):
        print(f"\nОбрабатываю {row_idx-1}/40: {company_name}")
        
        # Собираем контакты
        phone, email = scrape_website(website, company_name)
        
        # Обновляем ячейки в Excel
        ws.cell(row=row_idx, column=3, value=phone)  # Телефон
        ws.cell(row=row_idx, column=4, value=email)  # Email
        
        # Небольшая задержка между запросами
        time.sleep(1)
    
    # Сохраняем обновленный файл
    wb.save('companies_data_updated.xlsx')
    print("\n" + "=" * 50)
    print("✓ Excel файл обновлен: companies_data_updated.xlsx")

if __name__ == "__main__":
    update_excel_with_contacts() 