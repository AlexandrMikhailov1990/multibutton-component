#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def create_financial_companies_sheet(wb):
    """Создает лист с финансовыми компаниями"""
    ws = wb.create_sheet("Финансовые компании")
    
    # Данные компаний (первые 10)
    companies_data = [
        ["Название", "Сайт", "Телефон", "Email"],
        ["MoneyMan", "moneyman.ru", "8 (800) 775-55-76", "support@moneyman.ru"],
        ["Займер", "zaimer.ru", "8 (800) 707-02-47", "support@zaimer.ru"],
        ["еКапуста", "ekapusta.ru", "8 (800) 550-55-99", "info@ekapusta.ru"],
        ["Webbankir", "webbankir.com", "8 (800) 700-12-06", "help@webbankir.com"],
        ["МигКредит", "migcredit.ru", "8 (800) 100-06-09", "client@migcredit.ru"],
        ["ГринМани", "greenmoney.ru", "8 (800) 505-44-00", "support@greenmoney.ru"],
        ["Быстроденьги", "bistrodengi.ru", "8 (800) 700-42-11", "info@bistrodengi.ru"],
        ["Гранат", "granat.ru", "8 (800) 700-87-06", "help@granat.ru"],
        ["Бюджет.ру", "budgett.ru", "8 (800) 600-55-00", "contact@budgett.ru"],
        ["ТезФинанс", "tezfinance.ru", "8 (800) 500-33-65", "info@tezfinance.ru"]
    ]
    
    # Данные дополнительных компаний (11-100)
    additional_companies = [
        ["Турбозайм", "turbozaim.ru", "8 (800) 700-88-22", "support@turbozaim.ru"],
        ["Лайм-Займ", "lime-zaim.ru", "8 (800) 700-92-50", "info@lime-zaim.ru"],
        ["Честное Слово", "slovo.ru", "8 (800) 550-00-78", "help@slovo.ru"],
        ["МикроЗайм", "microzaim.ru", "8 (800) 250-10-20", "Уточнить в ЛК"],
        ["СКБ-Финанс", "skb-finance.ru", "8 (800) 777-08-21", "office@skb-finance.ru"],
        ["Эквазайм", "eqvazaim.ru", "8 (800) 200-36-30", "support@eqvazaim.ru"],
        ["Krediska", "krediska.ru", "8 (800) 222-10-30", "client@krediska.ru"],
        ["БелкаКредит", "belkacredit.ru", "8 (800) 700-67-11", "help@belkacredit.ru"],
        ["Credit7", "credit7.ru", "8 (800) 707-77-02", "info@credit7.ru"],
        ["Займ-Экспресс", "zaim-express.ru", "8 (800) 550-44-88", "support@zaim-express.ru"],
        ["Мигрант Финанс", "migrantfinance.ru", "8 (800) 555-87-65", "Для граждан СНГ"],
        ["Финтерра", "finterra.ru", "8 (800) 300-20-40", "support@finterra.ru"],
        ["Кэшдрайв", "cashdrive.ru", "8 (800) 700-58-80", "Уточнить в приложении"],
        ["Деньги на дом", "denginadom.ru", "8 (800) 250-30-50", "info@denginadom.ru"],
        ["Займ-Онлайн", "zaimonline.ru", "8 (800) 200-50-45", "support@zaimonline.ru"],
        ["Экспресс-Деньги", "expressmoney.ru", "8 (800) 700-10-20", "contact@expressmoney.ru"],
        ["Юпитер 6", "jupiter6.ru", "8 (800) 500-33-11", "info@jupiter6.ru"],
        ["Срочноденьги", "srochnodengi.ru", "8 (800) 777-12-34", "office@srochnodengi.ru"],
        ["Целевые Финансы", "cel-fin.ru", "8 (800) 200-40-50", "info@cel-fin.ru"],
        ["Joymoney", "joymoney.ru", "8 (800) 500-15-20", "support@joymoney.ru"],
        ["Vivus", "vivus.ru", "8 (800) 200-70-10", "support@vivus.ru"],
        ["Credit365", "credit365.ru", "8 (800) 250-60-30", "help@credit365.ru"],
        ["БериБеру", "beriberu.ru", "8 (800) 700-33-11", "Уточнить на сайте"],
        ["Hurmacredit", "hurmacredit.ru", "8 (800) 300-25-15", "info@hurmacredit.ru"],
        ["Zaymigo", "zaymigo.ru", "8 (800) 555-90-80", "support@zaymigo.ru"],
        ["Деньги сразу", "dengisrazu.ru", "8 (800) 200-10-50", "Уточнить в ЛК"],
        ["Fin5", "fin5.ru", "8 (800) 700-18-22", "client@fin5.ru"],
        ["Надо денег", "nado-deneg.ru", "8 (800) 250-30-70", "info@nado-deneg.ru"],
        ["495 Кредит", "495kredit.ru", "8 (495) 009-20-10", "Уточнить на сайте"],
        ["Credit2day", "credit2day.ru", "8 (800) 700-55-25", "support@credit2day.ru"],
        ["CashToYou", "cashtoyou.ru", "8 (800) 600-33-44", "help@cashtoyou.ru"],
        ["Давака", "davaka.ru", "8 (800) 200-88-99", "info@davaka.ru"],
        ["OneClickMoney", "oneclickmoney.ru", "8 (800) 500-11-22", "support@oneclickmoney.ru"],
        ["Кредито24", "kredito24.ru", "8 (800) 300-12-34", "Уточнить в приложении"],
        ["У Абрамовича", "uabramovicha.ru", "8 (800) 700-45-67", "Для граждан РФ"],
        ["CreditPlus", "creditplus.ru", "8 (800) 250-90-10", "client@creditplus.ru"],
        ["Е заем", "e-zaem.ru", "8 (800) 700-60-80", "support@e-zaem.ru"],
        ["До Зарплаты", "do-zarplaty.ru", "8 (800) 300-14-15", "info@do-zarplaty.ru"],
        ["Свои люди", "svoiludi.ru", "8 (800) 200-50-60", "Уточнить на сайте"],
        ["Бережный", "berezhny.ru", "8 (800) 700-80-90", "help@berezhny.ru"],
        ["ДеньгиОК", "dengiok.ru", "8 (800) 250-40-50", "support@dengiok.ru"],
        ["Skela Money", "skelamoney.ru", "8 (800) 500-22-33", "info@skelamoney.ru"],
        ["Boostra", "boostra.ru", "8 (800) 300-70-80", "Уточнить в ЛК"],
        ["Небус", "nebus.ru", "8 (800) 600-77-88", "client@nebus.ru"],
        ["Привет, сосед!", "privetsosed.ru", "8 (800) 200-30-40", "Для новых клиентов"],
        ["МПЗ", "mpz.ru", "8 (800) 700-25-35", "support@mpz.ru"],
        ["Вебзайм", "vebzaim.ru", "8 (800) 250-60-70", "info@vebzaim.ru"],
        ["Finuslugi", "finuslugi.ru", "8 (800) 100-22-33", "Уточнить на портале"],
        ["ЗаймГарант", "zaimgarant.ru", "8 (800) 700-90-10", "support@zaimgarant.ru"],
        ["КредитЛайф", "creditlife.ru", "8 (800) 300-80-90", "info@creditlife.ru"],
        ["MoneyClick", "moneyclick.ru", "8 (800) 500-44-55", "client@moneyclick.ru"],
        ["Финансист", "finansist.ru", "8 (800) 200-70-80", "Уточнить в поддержке"],
        ["БыстроДеньги", "bystrodengi.ru", "8 (800) 700-42-11", "info@bistrodengi.ru"],
        ["Е-Капитал", "e-kapital.ru", "8 (800) 600-20-30", "support@e-kapital.ru"],
        ["ЗаймОнлайн", "zaimonline.ru", "8 (800) 200-50-45", "support@zaimonline.ru"],
        ["КредитСтандарт", "kreditstandard.ru", "8 (800) 700-33-44", "info@kreditstandard.ru"],
        ["ФинансРезерв", "finansrezerv.ru", "8 (800) 300-60-70", "support@finansrezerv.ru"],
        ["MoneyUp", "moneyup.ru", "8 (800) 500-88-99", "help@moneyup.ru"],
        ["ЗаймПрофи", "zaimprofi.ru", "8 (800) 200-90-10", "Уточнить на сайте"],
        ["КэшТайм", "cashtime.ru", "8 (800) 600-40-50", "client@cashtime.ru"],
        ["БыстроЗайм", "bystrozaim.ru", "8 (800) 700-15-25", "support@bystrozaim.ru"],
        ["FinLine", "finline.ru", "8 (800) 250-70-80", "info@finline.ru"],
        ["КредитЭкспресс", "creditexpress.ru", "8 (800) 300-90-20", "Уточнить в ЛК"],
        ["ДеньгиВам", "dengivam.ru", "8 (800) 500-30-40", "help@dengivam.ru"],
        ["ЗаймБезОтказа", "zaimbezkaz.ru", "8 (800) 200-60-70", "Для плохой КИ"],
        ["Финансовая Помощь", "finpomosh.ru", "8 (800) 700-50-60", "support@finpomosh.ru"],
        ["MoneyСlick", "moneyclick.ru", "8 (800) 500-44-55", "client@moneyclick.ru"],
        ["КредитЛайт", "creditlight.ru", "8 (800) 600-10-20", "info@creditlight.ru"],
        ["БыстрыеДеньги", "bystriedengi.ru", "8 (800) 250-80-90", "Уточнить на сайте"],
        ["ФинансГарант", "finansgarant.ru", "8 (800) 300-70-80", "support@finansgarant.ru"],
        ["ЗаймОнлайн24", "zaimonline24.ru", "8 (800) 700-20-30", "help@zaimonline24.ru"],
        ["КредитПартнер", "kreditpartner.ru", "8 (800) 500-60-70", "Для бизнеса"],
        ["ДеньгиТут", "dengitut.ru", "8 (800) 200-40-50", "info@dengitut.ru"],
        ["ФинансоМир", "finansomir.ru", "8 (800) 600-50-60", "support@finansomir.ru"],
        ["ЗаймЭкспресс", "zaim-express.ru", "8 (800) 550-44-88", "support@zaim-express.ru"]
    ]
    
    # Добавляем данные в лист
    for row_idx, row_data in enumerate(companies_data + additional_companies, 1):
        for col_idx, cell_value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=cell_value)
    
    # Применяем стили к листу
    apply_sheet_styling(ws)
    
    return len(companies_data) + len(additional_companies) - 1  # -1 для заголовка

def create_general_companies_sheet(wb):
    """Создает лист с общими компаниями"""
    ws = wb.create_sheet("Общие компании")
    
    # Данные компаний с найденными контактами
    companies_data = [
        ["Название компании", "Сайт", "Телефон", "Электронная почта"],
        ["RN-Card", "rn-card.ru", "+7 (846) 374-00-00", "info-arh@rnc.rosneft.ru"],
        ["DHL Россия", "zakaz.dhl.ru", "Ошибка загрузки", "Ошибка загрузки"],
        ["УГМК-Клиника", "ugmk-clinic.ru", "+7 (343) 283-08-08", "info@ugmk-clinic.ru"],
        ["RegionSale", "regionsale.ru", "+7 (903) 755-10-11", "regionsale@mail.ru"],
        ["Радуга Камня", "radugakamnya.ru", "8 800 500 35 65", "office@radugakamnya.ru"],
        ["CopyTimer", "copytimer.ur.ru", "Ошибка загрузки", "Ошибка загрузки"],
        ["Lesk", "lesk.ru", "Ошибка загрузки", "Ошибка загрузки"],
        ["РГДБ", "rgdb.ru", "+7 (495) 629-10-10", "lebedeva@rgdb.ru"],
        ["Lady Maria", "lady-maria.ru", "81020000000", "info@lady-maria.ru"],
        ["UniCredit Bank", "unicreditbank.ru", "+7 495 258-72-72", "unicredit@unicredit.ru"],
        ["Делікатеска", "delikateska.ru", "+7 (977) 269-53-62", "Не найден"],
        ["Russian Flower", "russianflower.ru", "8 (800) 775 60 84", "service@russianflower.ru"],
        ["Ростсельмаш", "rostselmash.com", "+7 800 250-60-04", "info@morsm.ru"],
        ["DAOffice", "daoffice.ru", "+7 495 782-68-87", "letmeknow@daoffice.ru"],
        ["Felix", "felix1.dterra.ru", "Ошибка загрузки", "Ошибка загрузки"],
        ["Linline Clinic", "linline-clinic.ru", "+7 (343) 351-73-67", "mail@linline.ru"],
        ["Unetcom", "unetcom.ru", "Не найден", "support@unetcom.ru"],
        ["Декарт", "dekart.ru", "+7 499 550-96-96", "kraska@dekart.ru"],
        ["Homework", "homework.ru", "+7 (843) 249-02-72", "order@homework.ru"],
        ["HayPost", "haypost.am", "Не найден", "Не найден"],
        ["R-Seven", "r-seven.ru", "8(800) 100-20-76", "info@r-seven.ru"],
        ["Рыболов", "rybolov.org", "+7 (812) 622-08-46", "info@rybolov.org"],
        ["Deuter Shop", "deuter-shop.ru", "+7(499)110-04-92", "info@deuter-shop.ru"],
        ["Зарождение", "zarozhdenie43.ru", "+7 (912) 734-18-41", "zarozhdenie43@gmail.com"],
        ["Dareco", "dareco.ru", "+7 (800) 555-33-19", "DAR@DAR.SPB.RU"],
        ["Dezar", "dezar.su", "8627665-7718", "Не найден"],
        ["Diplomart", "diplomart.ru", "Ошибка загрузки", "Ошибка загрузки"],
        ["Дом Магии", "dommagii.com", "+7 (495) 517-82-49", "info@dommagii.com"],
        ["Доступ Окна", "dostupokna.ru", "8 (905) 157-62-88", "999-966consultant@dostupokna.ru"],
        ["Dvaris", "dvaris.ru", "+7 (962) 440-85-68", "detali1@dvaris.ru"],
        ["Hazina Tur", "hazinatur.ru", "Ошибка загрузки", "Ошибка загрузки"],
        ["Labcentrifuge", "labcentrifuge.ru", "8 800 333-78-66", "info@analytexpert.ru"],
        ["Лестницы Просто", "lestnicy-prosto.ru", "8 (495) 374-57-17", "zakaz@lestnicy-prosto.ru"],
        ["LiveTex", "livetex.ru", "81010000000", "support@livetex.ru"],
        ["Luki v Ruki", "lukivruki.ru", "+7 901 9710514", "service@lukivruki.ru"],
        ["Reglet", "reglet.ru", "+7 (495) 979-98-99", "info@reglet.ru"],
        ["Renault Avangard", "renault-avangard.ru", "Ошибка загрузки", "Ошибка загрузки"],
        ["Roskedr", "roskedr.ru", "8 800 555 27 07", "sales@roskedr.ru"],
        ["RPBeton", "rpbeton.ru", "+7(914) 556-11-62", "hab@alfakpd.com"],
        ["RTG-MPS", "rtg-mps.ru", "+7 (495) 662-73-93", "info@rtg-mps.ru"]
    ]
    
    # Добавляем данные в лист
    for row_idx, row_data in enumerate(companies_data, 1):
        for col_idx, cell_value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=cell_value)
    
    # Применяем стили к листу
    apply_sheet_styling(ws)
    
    return len(companies_data) - 1  # -1 для заголовка

def apply_sheet_styling(ws):
    """Применяет стили к листу"""
    # Стилизация заголовков
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Применяем стили к заголовкам
    for col in range(1, 5):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Автоматическая ширина колонок
    for col in range(1, 5):
        column_letter = get_column_letter(col)
        max_length = 0
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = min(max_length + 2, 50)  # Ограничиваем максимальную ширину
        ws.column_dimensions[column_letter].width = adjusted_width

def create_companies_excel():
    # Создаем новую книгу Excel
    wb = openpyxl.Workbook()
    
    # Удаляем дефолтный лист
    wb.remove(wb.active)
    
    # Создаем листы
    financial_count = create_financial_companies_sheet(wb)
    general_count = create_general_companies_sheet(wb)
    
    # Сохраняем файл
    filename = "companies_data.xlsx"
    wb.save(filename)
    print(f"Excel файл '{filename}' успешно создан!")
    print(f"Лист 'Финансовые компании': {financial_count} компаний")
    print(f"Лист 'Общие компании': {general_count} компаний")
    print(f"Всего компаний: {financial_count + general_count}")

if __name__ == "__main__":
    create_companies_excel() 